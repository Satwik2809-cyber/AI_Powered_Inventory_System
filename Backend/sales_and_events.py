from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from datetime import date, datetime, timedelta
from database import engine
from models import User, Product, ProductBatch, Sale, SaleItem, Event, EventStock, EventSale, EventSaleItem, EventDay, MonthlyCount, UserArea, EventSeller, DailySaleRequest, DailySaleItemRequest, EventCreateRequest
from auth import verify_password
from pydantic import BaseModel
import io
import openpyxl

def get_product_by_identity(session, name: str, category: str, rate: float):
    return session.exec(
        select(Product).where(
            Product.name == name,
            Product.category == category,
            Product.rate == rate,
            Product.is_active == True
        )
    ).first()
router = APIRouter(tags=["Sales & Events"])

def daily_sale_logic(
        session: Session,
        sale_data: DailySaleRequest
):
    try:
        total_amount = 0
        total_quantity = 0
        sale_items = []
        
        print(f"🔥 DEBUG: Processing Daily Sale for User ID {sale_data.user_id}")

        for item in sale_data.items:
            product = None
            if item.product_id:
                product = session.get(Product, item.product_id)
            
            if not product:
                product = get_product_by_identity(
                    session,
                    item.name,
                    item.category,
                    item.rate
                )

            if not product:
                raise HTTPException(
                    status_code=404,
                    detail=f"Product not found: {item.name} ({item.category})"
                )

            batches = session.exec(
                select(ProductBatch)
                .where(ProductBatch.product_id == product.id)
                .order_by(ProductBatch.expiry_date)
            ).all()

            remaining_qty = item.quantity
            for batch in batches:
                if remaining_qty <= 0:
                    break
                if batch.quantity >= remaining_qty:
                    batch.quantity -= remaining_qty
                    remaining_qty = 0
                else:
                    remaining_qty -= batch.quantity
                    batch.quantity = 0

            if remaining_qty > 0:
                raise HTTPException(
                    status_code=400,
                    detail=f"Insufficient stock for {item.name}. Missing: {remaining_qty}"
                )

            item_total = 0 if getattr(item, "is_gift", False) else (item.quantity * item.rate)
            total_amount += item_total
            total_quantity += item.quantity
            
            sale_items.append(
                SaleItem(
                    product_id=product.id,
                    quantity=item.quantity,
                    rate=item.rate,
                    total_price=item_total,
                    is_gift=getattr(item, "is_gift", False)
                )
            )

        sale = Sale(
            user_id=sale_data.user_id,
            sale_type="daily",
            total_amount=total_amount,
            total_quantity=total_quantity,
            payment_mode=sale_data.payment_mode
        )
        session.add(sale)
        session.commit()
        session.refresh(sale)

        for si in sale_items:
            si.sale_id = sale.id
            session.add(si)

        session.commit()

        return {
            "message": "Sale completed successfully",
            "sale_id": sale.id,
            "total_amount": total_amount,
            "total_quantity": total_quantity
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ ERROR in daily_sale_logic: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/sales/daily")
def daily_sale(
    sale_data: DailySaleRequest
):
    with Session(engine) as session:
        return daily_sale_logic(session, sale_data)


@router.post("/events")
def create_event(req: EventCreateRequest):
    is_multi_day = (req.mode == "multi-day")
    total_days = req.days if req.days else 1
    if is_multi_day and total_days < 2:
        raise HTTPException(
            status_code=400,
            detail="Multi-day event must have at least 2 days"
        )
    selling_mode = "open_sell" if req.open_sell else "area_wise"
    try:
        start_date = datetime.strptime(req.start_date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid start_date format, expected YYYY-MM-DD")

    event = Event(
       name=req.name,
       is_multi_day=is_multi_day,
       days=total_days,
       selling_mode=selling_mode,
       status="created",
       started_at=start_date
    )
    with Session(engine) as session:
        session.add(event)
        session.commit()
        session.refresh(event)
        return event

class EventStatusUpdate(BaseModel):
    status: str

@router.put("/events/{event_id}/status")
def update_event_status(event_id: int, req: EventStatusUpdate):
    with Session(engine) as session:
        event = session.get(Event, event_id)
        if not event:
            raise HTTPException(404, "Event not found")
        event.status = req.status
        session.commit()
        return {"message": "Status updated", "status": event.status}

@router.get("/events/live-dashboard/{event_id}")
def get_live_dashboard(event_id: int):
    # Placeholder for live dashboard data
    return {"sales": 0, "revenue": 0}

@router.get("/events")
def get_all_events():
    with Session(engine) as session:
        events = session.exec(select(Event).order_by(Event.created_at.desc())).all()
        result = []
        for e in events:
            sellers = session.exec(
                select(User).join(EventSeller, EventSeller.user_id == User.id)
                .where(EventSeller.event_id == e.id)
            ).all()
            assigned_sellers = [{"id": s.id, "name": s.name or s.username} for s in sellers]

            result.append({
                "id": e.id,
                "name": e.name,
                "status": e.status,
                "mode": "multi-day" if e.is_multi_day else "single-day",
                "open_sell": e.selling_mode == "open_sell",
                "start_date": e.started_at.strftime("%Y-%m-%d") if e.started_at else "",
                "days": e.days,
                "assigned_sellers": assigned_sellers
            })
        return result


class EventSellerAssignment(BaseModel):
    event_id: int
    seller_ids: list[int]

@router.post("/events/assign-sellers")
def assign_event_sellers(req: EventSellerAssignment):
    with Session(engine) as session:

        event = session.get(Event, req.event_id)

        if not event:
            raise HTTPException(status_code=404, detail="Event not found")

        # Remove old sellers
        old = session.exec(
            select(EventSeller).where(EventSeller.event_id == event.id)
        ).all()

        for s in old:
            session.delete(s)

        # Add new sellers
        for uid in req.seller_ids:
            session.add(EventSeller(event_id=event.id, user_id=uid))

        session.commit()

        return {
            "message": "Event sellers assigned",
            "event": event.name,
            "users": req.seller_ids
        }

from models import User, Product, ProductBatch, Sale, SaleItem, Event, EventStock, EventStockHistory, EventSale, EventSaleItem, EventDay, MonthlyCount, UserArea, EventSeller, DailySaleRequest, DailySaleItemRequest, EventCreateRequest, EventSellRequest

def pack_items_logic(session:Session,
                     event_name: str,
                     user_id: int,
                     items: list[dict]):
    # 1️⃣ Find Event
        event = session.exec(
            select(Event).where(Event.name == event_name)
        ).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        if event.status not in ["created", "active", "packing"]:
            raise HTTPException(
                status_code=400,
                detail="Cannot pack items for this event state"
            )
        # 2️⃣ Process each item
        for item in items:
            product = get_product_by_identity(
                session,
                item["name"],
                item["category"],
                item["rate"]
            )
            if not product:
                raise HTTPException(
                    status_code=404,
                    detail=f"Product not found: {item['name']}"
                )
            required_qty = item["quantity"]
            # 3️⃣ Get Main Vault batches (FIFO by expiry)
            batches = session.exec(
                select(ProductBatch)
                .where(ProductBatch.product_id == product.id)
                .order_by(ProductBatch.expiry_date)
            ).all()
            remaining_to_pack = required_qty
            for batch in batches:
                if remaining_to_pack <= 0:
                    break
                if batch.quantity >= remaining_to_pack:
                    batch.quantity -= remaining_to_pack
                    remaining_to_pack = 0
                else:
                    remaining_to_pack -= batch.quantity
                    batch.quantity = 0

            if remaining_to_pack > 0:
                raise HTTPException(
                    status_code=400,
                    detail=f"Insufficient stock for {item['name']}"
                )
            
            # 4️⃣ Add / Update Event Stock
            is_gift = item.get("is_gift", False)
            event_stock = session.exec(
                select(EventStock).where(
                    EventStock.event_id == event.id,
                    EventStock.product_id == product.id,
                    EventStock.is_gift == is_gift
                )
            ).first()
            if event_stock:
                event_stock.quantity_taken += required_qty
                event_stock.quantity_remaining += required_qty
            else:
                event_stock = EventStock(
                    event_id=event.id,
                    product_id=product.id,
                    quantity_taken=required_qty,
                    quantity_remaining=required_qty,
                    is_gift=is_gift
                )
                session.add(event_stock)
                
            # 5️⃣ Log Packing History
            action_name = "restocked" if event.status == "active" else "packed"
            history_entry = EventStockHistory(
                event_id=event.id,
                product_id=product.id,
                user_id=user_id,
                quantity=required_qty,
                action=action_name
            )
            session.add(history_entry)

        session.commit()
        return {
            "message": "Items packed successfully for event",
            "event": event.name
        }

@router.post("/events/pack")
def pack_items_for_event(
    event_name: str,
    user_id: int,
    items: list[dict]
):
    with Session(engine) as session:
        return pack_items_logic(session, event_name, user_id, items)

def event_sell_logic(session:Session, req: EventSellRequest):
    # 1️⃣ Get Event
        event = session.exec(
            select(Event).where(Event.name == req.event_name)
        ).first()

        if not event or event.status != "active":
            raise HTTPException(status_code=400, detail="Event not active")

        # 🔐 OPEN SELL: check seller assignment
        if event.selling_mode == "open_sell":
            allowed = session.exec(
                select(EventSeller).where(
                    EventSeller.event_id == event.id,
                    EventSeller.user_id == req.user_id
                )
            ).first()

            if not allowed:
                raise HTTPException(
                    status_code=403,
                    detail="You are not assigned as seller for this event"
                )

        # 🔒 AREA-WISE: fetch user areas
        if event.selling_mode == "area_wise":
            user_areas = session.exec(
                select(UserArea.category)
                .where(UserArea.user_id == req.user_id)
            ).all()

        # 2️⃣ Check / create event day
        day = session.exec(
            select(EventDay).where(
                EventDay.event_id == event.id,
                EventDay.day_number == req.day_number
            )
        ).first()

        if not day:
            day = EventDay(
                event_id=event.id,
                day_number=req.day_number,
                status="active"
            )
            session.add(day)
            session.commit()

        total_amount = 0
        sale_items = []

        # 3️⃣ Process items
        for item in req.items:
            product = get_product_by_identity(
                session,
                item.name,
                item.category,
                item.rate
            )

            if not product:
                raise HTTPException(
                    status_code=404,
                    detail=f"Product not found: {item.name}"
                )

            # 🔒 AREA CHECK ONLY IF AREA_WISE
            if event.selling_mode == "area_wise":
                if product.category not in user_areas:
                    raise HTTPException(
                        status_code=403,
                        detail=f"You are not allowed to sell {product.name}"
                    )

            # 4️⃣ Check event stock
            event_stock = session.exec(
                select(EventStock).where(
                    EventStock.event_id == event.id,
                    EventStock.product_id == product.id
                )
            ).first()

            if not event_stock or event_stock.quantity_remaining < item.quantity:
                raise HTTPException(
                    status_code=400,
                    detail=f"Insufficient event stock for {item.name}"
                )

            # 5️⃣ Deduct event stock
            event_stock.quantity_remaining -= item.quantity

            item_total = 0 if getattr(item, "is_gift", False) else (item.quantity * item.rate)
            total_amount += item_total

            sale_items.append(
                EventSaleItem(
                    product_id=product.id,
                    quantity=item.quantity,
                    rate=item.rate,
                    total_price=item_total,
                    is_gift=getattr(item, "is_gift", False)
                )
            )

        # 6️⃣ Create Event Sale
        sale = EventSale(
            event_id=event.id,
            day_number=req.day_number,
            user_id=req.user_id,
            payment_mode=req.payment_mode,
            total_amount=total_amount
        )

        session.add(sale)
        session.commit()
        session.refresh(sale)

        # 7️⃣ Save sale items
        for si in sale_items:
            si.event_sale_id = sale.id
            session.add(si)

        session.commit()

        return {
            "message": "Event sale completed",
            "event": event.name,
            "day": req.day_number,
            "total_amount": total_amount,
            "selling_mode": event.selling_mode
        }

@router.post("/events/sell")
def event_sell(req: EventSellRequest):
    with Session(engine) as session:
        return event_sell_logic(session, req)
    
@router.post("/events/day/close")
def close_event_day(event_name: str, day_number: int):
    with Session(engine) as session:
        event = session.exec(
            select(Event).where(Event.name == event_name)
        ).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        day = session.exec(
            select(EventDay).where(
                EventDay.event_id == event.id,
                EventDay.day_number == day_number
            )
        ).first()
        if not day:
            raise HTTPException(status_code=404, detail="Event day not found")

        if day.status == "closed":
            return {"message": "Day already closed"}
        day.status = "closed"
        day.ended_at = datetime.utcnow()
        session.commit()
        return {
            "message": "Event day closed successfully",
            "event": event.name,
            "day": day_number
        }
    
@router.get("/events/day/summary/user")
def user_day_summary(event_name: str, day_number: int, user_id: int):
    with Session(engine) as session:
        event = session.exec(
            select(Event).where(Event.name == event_name)
        ).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        sales = session.exec(
            select(EventSale).where(
                EventSale.event_id == event.id,
                EventSale.day_number == day_number,
                EventSale.user_id == user_id
            )
        ).all()
        cash = 0
        online = 0
        item_summary = {}
        for sale in sales:
            if sale.payment_mode == "cash":
                cash += sale.total_amount
            else:
                online += sale.total_amount
            items = session.exec(
                select(EventSaleItem).where(
                    EventSaleItem.event_sale_id == sale.id
                )
            ).all()
            for item in items:
                item_summary[item.product_id] = (
                    item_summary.get(item.product_id, 0) + item.quantity
                )
        return {
            "event": event.name,
            "day": day_number,
            "user_id": user_id,
            "cash_total": cash,
            "online_total": online,
            "grand_total": cash + online,
            "items_sold": item_summary
        }
    
@router.get("/events/day/summary/admin")
def admin_day_summary(event_name: str, day_number: int):
    with Session(engine) as session:
        event = session.exec(
            select(Event).where(Event.name == event_name)
        ).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        sales = session.exec(
            select(EventSale).where(
                EventSale.event_id == event.id,
                EventSale.day_number == day_number
            )
        ).all()
        cash = 0
        online = 0
        user_summary = {}
        item_summary = {}
        detailed_sales = []
        for sale in sales:
            if sale.payment_mode == "cash":
                cash += sale.total_amount
            else:
                online += sale.total_amount
                
            # user info
            user = session.get(User, sale.user_id)
            username = user.username if user else f"User {sale.user_id}"
            if username not in user_summary:
                user_summary[username] = {"total_amount": 0, "sales_count": 0}
            user_summary[username]["total_amount"] += sale.total_amount
            user_summary[username]["sales_count"] += 1
            
            # item info
            items = session.exec(
                select(EventSaleItem, Product)
                .join(Product, EventSaleItem.product_id == Product.id)
                .where(EventSaleItem.event_sale_id == sale.id)
            ).all()
            
            sale_items_detail = []
            for si, p in items:
                if p.id not in item_summary:
                    item_summary[p.id] = {
                        "product_name": p.name,
                        "category": p.category,
                        "quantity_sold": 0,
                        "rate": p.rate,
                        "total_price": 0
                    }
                item_summary[p.id]["quantity_sold"] += si.quantity
                item_summary[p.id]["total_price"] += (si.quantity * si.rate)
                
                sale_items_detail.append({
                    "name": p.name,
                    "quantity": si.quantity,
                    "rate": si.rate,
                    "total": si.quantity * si.rate
                })
                
            detailed_sales.append({
                "id": sale.id,
                "timestamp": sale.timestamp.isoformat() if hasattr(sale, "timestamp") and sale.timestamp else "",
                "payment_mode": sale.payment_mode,
                "total_amount": sale.total_amount,
                "seller_name": username,
                "items": sale_items_detail
            })
                
        stock_summary_list = list(item_summary.values())

        return {
            "event": event.name,
            "day": day_number,
            "sales": len(sales),
            "cash_total": cash,
            "online_total": online,
            "grand_total": cash + online,
            "user_wise_total": user_summary,
            "stock_summary": stock_summary_list,
            "detailed_sales": detailed_sales
        }

@router.post("/events/close")
def close_event(event_name: str):
    with Session(engine) as session:
        event = session.exec(
            select(Event).where(Event.name == event_name)
        ).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        if event.status == "completed":
            return {"message": "Event already completed"}
        event.status = "closed_pending_return"
        session.commit()
        return {
            "message": "Event closed. Please verify remaining stock before return.",
            "event": event.name,
            "status": event.status
        }
@router.get("/events/remaining-stock")
def view_remaining_event_stock(event_name: str):
    with Session(engine) as session:
        event = session.exec(
            select(Event).where(Event.name == event_name)
        ).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        stocks = session.exec(
            select(EventStock, Product)
            .join(Product, EventStock.product_id == Product.id)
            .where(EventStock.event_id == event.id)
        ).all()
        
        result = []
        for stock, product in stocks:
            result.append({
                "product_id": product.id,
                "name": product.name,
                "category": product.category,
                "rate": product.rate,
                "quantity_taken": stock.quantity_taken,
                "quantity_remaining": stock.quantity_remaining,
                "is_gift": stock.is_gift,
                "checked_in": stock.checked_in,
                "checked_out": stock.checked_out
            })
            
        return {
            "event": event.name,
            "remaining_stock": result
        }

class EventItemUpdateRequest(BaseModel):
    checked_in: bool = None
    checked_out: bool = None
    quantity_taken: int = None

@router.patch("/events/{event_name}/items/{product_id}")
def update_event_item(event_name: str, product_id: int, req: EventItemUpdateRequest):
    with Session(engine) as session:
        event = session.exec(select(Event).where(Event.name == event_name)).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        
        stock = session.exec(
            select(EventStock)
            .where(EventStock.event_id == event.id, EventStock.product_id == product_id)
        ).first()
        if not stock:
            raise HTTPException(status_code=404, detail="Item not found in event")

        if req.checked_in is not None:
            stock.checked_in = req.checked_in
        if req.checked_out is not None:
            stock.checked_out = req.checked_out
        if req.quantity_taken is not None:
            # We need to compute the difference to update quantity_remaining as well.
            diff = req.quantity_taken - stock.quantity_taken
            stock.quantity_taken = req.quantity_taken
            stock.quantity_remaining += diff

            # Update Main Vault ProductBatch as well
            product_batch = session.exec(
                select(ProductBatch)
                .where(ProductBatch.product_id == product_id)
                .order_by(ProductBatch.manufacturing_date)
            ).first()
            if product_batch:
                product_batch.quantity -= diff
            else:
                # If there's no batch but quantity changed, we should just assume they deal with it or log it
                pass

        session.commit()
        return {"message": "Event item updated successfully"}


@router.get("/events/stock-history")
def get_event_stock_history(event_name: str):
    with Session(engine) as session:
        event = session.exec(
            select(Event).where(Event.name == event_name)
        ).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
            
        history = session.exec(
            select(EventStockHistory, Product, User)
            .join(Product, EventStockHistory.product_id == Product.id)
            .join(User, EventStockHistory.user_id == User.id)
            .where(EventStockHistory.event_id == event.id)
            .order_by(EventStockHistory.created_at.desc())
        ).all()
        
        result = []
        for h, p, u in history:
            result.append({
                "id": h.id,
                "product_name": p.name,
                "quantity": h.quantity,
                "action": h.action,
                "packed_by": u.username,
                "created_at": h.created_at.isoformat()
            })
            
        return {
            "event": event.name,
            "history": result
        }

@router.get("/events/summary/{event_name}")
def get_full_event_summary(event_name: str):
    with Session(engine) as session:
        event = session.exec(
            select(Event).where(Event.name == event_name)
        ).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
            
        # Get all sales for this event
        sales = session.exec(
            select(EventSale).where(EventSale.event_id == event.id)
        ).all()
        
        # Get stock details for items left/sold
        event_stocks = session.exec(
            select(EventStock, Product)
            .join(Product, EventStock.product_id == Product.id)
            .where(EventStock.event_id == event.id)
        ).all()
        
        stock_summary = []
        for stock, product in event_stocks:
            stock_summary.append({
                "product_name": product.name,
                "category": product.category,
                "quantity_taken": stock.quantity_taken,
                "quantity_sold": stock.quantity_taken - stock.quantity_remaining,
                "quantity_remaining": stock.quantity_remaining,
                "rate": product.rate
            })
        
        grand_total = 0
        total_cash = 0
        total_online = 0
        day_wise = {}
        seller_totals = {}
        
        for sale in sales:
            amount = sale.total_amount
            grand_total += amount
            
            if sale.payment_mode == "cash":
                total_cash += amount
            else:
                total_online += amount
                
            # Day grouping
            day = sale.day_number
            if day not in day_wise:
                day_wise[day] = {"cash": 0, "online": 0, "total": 0}
            
            if sale.payment_mode == "cash":
                day_wise[day]["cash"] += amount
            else:
                day_wise[day]["online"] += amount
            day_wise[day]["total"] += amount
                
            # Seller grouping
            user = session.get(User, sale.user_id)
            username = user.username if user else f"User {sale.user_id}"
            if username not in seller_totals:
                seller_totals[username] = 0
            seller_totals[username] += amount

        # Assigned sellers
        sellers_assigned = []
        if event.selling_mode == "open_sell":
             sellers = session.exec(
                select(User).join(EventSeller, EventSeller.user_id == User.id)
                .where(EventSeller.event_id == event.id)
             ).all()
             sellers_assigned = [u.username for u in sellers]
             
        # Recent restocks logic
        recent_restocks = []
        if event.started_at:
            hist_records = session.exec(
                select(EventStockHistory)
                .where(EventStockHistory.event_id == event.id, EventStockHistory.action == "packed", EventStockHistory.created_at > event.started_at)
                .order_by(EventStockHistory.created_at.desc())
            ).all()
            for h in hist_records:
                p = session.get(Product, h.product_id)
                if p:
                    u = session.get(User, h.user_id) if h.user_id else None
                    recent_restocks.append({
                        "product_name": p.name,
                        "quantity": h.quantity,
                        "date": h.created_at.isoformat(),
                        "packed_by": u.username if u else "System"
                    })

        return {
            "event_name": event.name,
            "status": event.status,
            "mode": "multi-day" if event.is_multi_day else "single-day",
            "selling_mode": event.selling_mode,
            "assigned_sellers": sellers_assigned,
            "financials": {
                "grand_total": grand_total,
                "total_cash": total_cash,
                "total_online": total_online,
            },
            "day_wise": day_wise,
            "seller_totals": seller_totals,
            "stock_summary": stock_summary,
            "recent_restocks": recent_restocks
        }


@router.get("/events/summary/excel/{event_name}")
def download_event_excel_summary(event_name: str):
    with Session(engine) as session:
        event = session.exec(
            select(Event).where(Event.name == event_name)
        ).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        
        # Determine total days and create a mapping for day-wise tracking
        is_multi_day = event.is_multi_day
        # Find the maximum day number sold or use event.days
        max_day = event.days if hasattr(event, "days") and event.days else 1
        
        # Get all sales for this event
        sales = session.exec(
            select(EventSale).where(EventSale.event_id == event.id)
        ).all()
        
        # Dictionary to track sale items:
        # product_id -> { "day_1": qty, "day_2": qty ... }
        daily_sales_map = {}
        for sale in sales:
            if sale.day_number > max_day:
                max_day = sale.day_number # Keep dynamic if sales ran over
                
            items = session.exec(
                select(EventSaleItem).where(EventSaleItem.event_sale_id == sale.id)
            ).all()
            
            for item in items:
                if item.product_id not in daily_sales_map:
                    daily_sales_map[item.product_id] = {}
                day_key = f"day_{sale.day_number}"
                daily_sales_map[item.product_id][day_key] = daily_sales_map[item.product_id].get(day_key, 0) + item.quantity
        
        # Get stock details for items packed
        event_stocks = session.exec(
            select(EventStock, Product)
            .join(Product, EventStock.product_id == Product.id)
            .where(EventStock.event_id == event.id)
        ).all()

        stock_history = session.exec(
            select(EventStockHistory)
            .where(EventStockHistory.event_id == event.id, EventStockHistory.action == "packed")
            .order_by(EventStockHistory.created_at)
        ).all()
        
        history_map = {}
        for h in stock_history:
            if h.product_id not in history_map:
                history_map[h.product_id] = []
            history_map[h.product_id].append((h.quantity, h.created_at))

        wb = openpyxl.Workbook()
        
        # Sort event_stocks by category and name for grouping
        event_stocks = sorted(event_stocks, key=lambda x: (x[1].category or "Uncategorized", x[1].name))

        # ==========================================
        # SHEET 1: OVERALL SUMMARY
        # ==========================================
        overall_ws = wb.active
        overall_ws.title = "Summary"
        headers = ["Item Name", "Rate", "Total Quantity Packed", "Restocked After Start", "Balance", "Sell", "Amount"]
        overall_ws.append(headers)

        # Style headers
        for cell in overall_ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            
        grand_total_sale = 0
        current_cat_summary = None
            
        for stock, product in event_stocks:
            cat = product.category or "Uncategorized"
            if cat != current_cat_summary:
                current_cat_summary = cat
                overall_ws.append([f"--- {cat.upper()} ---", "", "", "", "", "", ""])
                for col_idx in range(1, len(headers)+1):
                    overall_ws.cell(row=overall_ws.max_row, column=col_idx).font = openpyxl.styles.Font(bold=True, italic=True)
                    overall_ws.cell(row=overall_ws.max_row, column=col_idx).fill = openpyxl.styles.PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")

            qty_taken = stock.quantity_taken
            qty_remaining = stock.quantity_remaining
            qty_sold = qty_taken - qty_remaining
            amount = qty_sold * product.rate
            grand_total_sale += amount
            
            restock_str = ""
            if product.id in history_map:
                entries = history_map[product.id]
                restock_texts = []
                for h_qty, dt in entries:
                    if event.started_at and dt > event.started_at:
                        restock_texts.append(f"+{h_qty} on {dt.strftime('%b %d, %H:%M')}")
                if restock_texts:
                    restock_str = " | ".join(restock_texts)

            balance = qty_remaining
            sell = qty_sold if qty_sold > 0 else "-"
            amt = amount if amount > 0 else "-"
            
            overall_ws.append([
                product.name,
                product.rate,
                qty_taken,
                restock_str,
                balance,
                sell,
                amt
            ])

        # Add grand total row
        overall_ws.append(["", "", "", "", "", "GRAND TOTAL:", grand_total_sale])
        
        # Add Cash/Online breakdown
        total_cash = sum(s.total_amount for s in sales if s.payment_mode == 'cash')
        total_online = sum(s.total_amount for s in sales if s.payment_mode != 'cash')
        overall_ws.append(["", "", "", "", "", "CASH:", total_cash])
        overall_ws.append(["", "", "", "", "", "ONLINE:", total_online])
        
        # Adjust column widths
        for col in overall_ws.columns:
            max_length = 0
            column = col[0].column_letter 
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            overall_ws.column_dimensions[column].width = adjusted_width

        # ==========================================
        # MULTI-DAY SHEETS
        # ==========================================
        if is_multi_day:
            for day in range(1, max_day + 1):
                day_ws = wb.create_sheet(title=f"Day {day}")
                day_ws.append(headers)
                
                for cell in day_ws[1]:
                    cell.font = openpyxl.styles.Font(bold=True)
                
                day_grand_total = 0
                current_cat_day = None
                
                for stock, product in event_stocks:
                    cat = product.category or "Uncategorized"
                    if cat != current_cat_day:
                        current_cat_day = cat
                        day_ws.append([f"--- {cat.upper()} ---", "", "", "", "", "", ""])
                        for col_idx in range(1, len(headers)+1):
                            day_ws.cell(row=day_ws.max_row, column=col_idx).font = openpyxl.styles.Font(bold=True, italic=True)
                            day_ws.cell(row=day_ws.max_row, column=col_idx).fill = openpyxl.styles.PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")

                    day_key = f"day_{day}"
                    
                    # Quantity sold on this day specifically
                    qty_sold_today = daily_sales_map.get(product.id, {}).get(day_key, 0)
                    
                    # Calculate opening balance: Total packed minus sales from all PREVIOUS days
                    sales_prior_to_today = 0
                    for prev_d in range(1, day):
                        sales_prior_to_today += daily_sales_map.get(product.id, {}).get(f"day_{prev_d}", 0)
                    
                    qty_opening_today = stock.quantity_taken - sales_prior_to_today
                    qty_balance_end_of_day = qty_opening_today - qty_sold_today
                    
                    amount_today = qty_sold_today * product.rate
                    day_grand_total += amount_today
                    
                    # We output ALL products packed for the event, even if 0 sold today, 
                    # so the columns exist and are blank/0 as requested.
                    day_ws.append([
                        product.name,
                        product.rate,
                        qty_opening_today,
                        restock_str,
                        qty_balance_end_of_day,
                        qty_sold_today,
                        amount_today
                    ])
                    
                day_ws.append(["", "", "", "", "", f"DAY {day} TOTAL:", day_grand_total])
                
                # Adjust column widths
                for col in day_ws.columns:
                    max_length = 0
                    column = col[0].column_letter 
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    day_ws.column_dimensions[column].width = adjusted_width

        # Write to byte stream
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        
        safe_name = event_name.replace(" ", "_")
        filename = f"Event_Summary_{safe_name}.xlsx"

        return StreamingResponse(
            iter([stream.getvalue()]), 
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

class EventUpdateRequest(BaseModel):
    name: str = None
    start_date: str = None
    days: int = None
    mode: str = None

@router.put("/events/{event_id}/details")
def edit_event_details(event_id: int, req: EventUpdateRequest):
    with Session(engine) as session:
        event = session.get(Event, event_id)
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        
        if req.name is not None:
            # Check if name exists
            existing = session.exec(select(Event).where(Event.name == req.name)).first()
            if existing and existing.id != event_id:
                raise HTTPException(status_code=400, detail="Event name already taken")
            event.name = req.name
            
        if req.start_date is not None:
            event.started_at = datetime.strptime(req.start_date, "%Y-%m-%d")
        if req.days is not None:
            event.days = req.days
        if req.mode is not None:
            event.is_multi_day = (req.mode == "multi-day")

        session.commit()
        return {"message": "Event details updated successfully"}

@router.get("/events/{event_id}")
def get_event(event_id: int):
    with Session(engine) as session:
        e = session.get(Event, event_id)
        if not e:
            raise HTTPException(status_code=404, detail="Event not found")
        
        sellers = session.exec(
            select(User).join(EventSeller, EventSeller.user_id == User.id)
            .where(EventSeller.event_id == e.id)
        ).all()
        assigned_sellers = [{"id": s.id, "name": s.name or s.username} for s in sellers]

        return {
            "id": e.id,
            "name": e.name,
            "status": e.status,
            "mode": "multi-day" if e.is_multi_day else "single-day",
            "open_sell": e.selling_mode == "open_sell",
            "start_date": e.started_at.strftime("%Y-%m-%d") if e.started_at else "",
            "days": e.days,
            "assigned_sellers": assigned_sellers
        }

@router.put("/events/{event_id}/status")
def update_event_status(event_id: int, payload: dict):
    with Session(engine) as session:
        e = session.get(Event, event_id)
        if not e:
            raise HTTPException(status_code=404, detail="Event not found")
            
        status = payload.get("status")
        if not status:
            raise HTTPException(status_code=400, detail="Status is required")
            
        e.status = status
        
        # If completing, set completed_at
        if status in ["completed", "Completed", "closed"]:
            from datetime import datetime
            e.completed_at = datetime.utcnow()
            
        session.commit()
        return {"message": "Status updated", "status": e.status}

@router.post("/events/return-stock")
def return_event_stock_to_main(event_name: str):
    with Session(engine) as session:
        event = session.exec(
            select(Event).where(Event.name == event_name)
        ).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        if event.status == "completed":
            raise HTTPException(
                status_code=400,
                detail="Event is already completed and stock has been returned."
            )
        try:
            event_stocks = session.exec(
                select(EventStock).where(EventStock.event_id == event.id)
            ).all()
            for es in event_stocks:
                if es.quantity_remaining > 0:
                    batch = ProductBatch(
                        product_id=es.product_id,
                        quantity=int(es.quantity_remaining),
                        manufacturing_date=None,
                        expiry_date=None,
                        last_restocked_at=datetime.utcnow()
                    )
                    session.add(batch)
                    
                    # Log the return in StockHistory
                    from models import StockHistory
                    history = StockHistory(
                        product_id=es.product_id,
                        quantity=int(es.quantity_remaining),
                        restocked_by="System (Event Return)"
                    )
                    session.add(history)
                    
                    es.quantity_remaining = 0
            event.status = "completed"
            session.commit()
            return {
                "message": "Remaining stock returned to Main Vault successfully",
                "event": event.name,
                "status": "completed"
            }
        except Exception as e:
            session.rollback()
            print(f"❌ ERROR in return_event_stock_to_main: {str(e)}")
            import traceback
            traceback.print_exc()
            raise HTTPException(status_code=500, detail=str(e))

class MonthlyPreviewRequest(BaseModel):
    month: str # "YYYY-MM"

@router.post("/monthly/preview")
def preview_monthly_count(data: MonthlyPreviewRequest):
    with Session(engine) as session:
        try:
            # Parse month string "YYYY-MM"
            year, month = map(int, data.month.split("-"))
            start_date = datetime(year, month, 1)
            
            # Calculate end_date (first day of next month)
            if month == 12:
                end_date = datetime(year + 1, 1, 1)
            else:
                end_date = datetime(year, month + 1, 1)
                
        except ValueError:
            raise HTTPException(400, "Invalid month format. Use YYYY-MM")

        # 1. Daily Sales in range
        daily_sales = session.exec(
            select(Sale).where(
                Sale.created_at >= start_date,
                Sale.created_at < end_date
            )
        ).all()

        # 2. Event Sales in range
        event_sales = session.exec(
            select(EventSale).where(
                EventSale.created_at >= start_date,
                EventSale.created_at < end_date
            )
        ).all()

        cash = 0
        online = 0

        for sale in daily_sales + event_sales:
            if sale.payment_mode == "cash":
                cash += sale.total_amount
            elif sale.payment_mode == "online":
                online += sale.total_amount
            else:
                 # Fallback if payment_mode is missing or different
                 cash += sale.total_amount

        return {
            "message": "Monthly count preview created",
            "period": f"{start_date.date()} → {end_date.date()}",
            "period_iso": data.month,
            "revenue": cash + online, # matches 'revenue' in frontend interface Preview
            "sales": len(daily_sales) + len(event_sales),
            "stock_value": 0, # Placeholder or calculate if needed
            "profit": 0,      # Placeholder
            "cash": cash,
            "online": online,
            "grand_total": cash + online,
            "status": "draft"
        }
@router.post("/monthly/confirm")
def confirm_monthly_count(data: MonthlyPreviewRequest):
    with Session(engine) as session:
        # 1. Start/End Dates
        try:
            year, month = map(int, data.month.split("-"))
            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1)
            else:
                end_date = datetime(year, month + 1, 1)
        except ValueError:
             raise HTTPException(400, "Invalid month")

        # 2. Check if already confirmed
        existing = session.exec(
            select(MonthlyCount).where(MonthlyCount.start_date == start_date)
            .where(MonthlyCount.status == "confirmed")
        ).first()
        
        if existing:
             return {"message": "Month already confirmed"}

        # 3. Calculate Totals
        daily_sales = session.exec(
            select(Sale).where(
                Sale.created_at >= start_date,
                Sale.created_at < end_date
            )
        ).all()
        event_sales = session.exec(
            select(EventSale).where(
                EventSale.created_at >= start_date,
                EventSale.created_at < end_date
            )
        ).all()

        cash = 0
        online = 0
        for sale in daily_sales + event_sales:
            if sale.payment_mode == "cash":
                cash += sale.total_amount
            elif sale.payment_mode == "online":
                 online += sale.total_amount
            else:
                 cash += sale.total_amount

        # 4. Save Record
        record = MonthlyCount(
            start_date=start_date,
            end_date=end_date,
            total_cash=cash,
            total_online=online,
            grand_total=cash + online,
            created_by="admin", # Placeholder or fetch user
            status="confirmed"
        )
        session.add(record)
        session.commit()

        return {
            "message": "Monthly count confirmed and locked",
            "period": f"{data.month}",
        }
@router.get("/monthly/history")
def monthly_history():
    with Session(engine) as session:
        return session.exec(
            select(MonthlyCount).order_by(MonthlyCount.created_at.desc())
        ).all()

@router.get("/events/stock/user")
def get_user_event_stock(event_name: str, user_id: int):
    with Session(engine) as session:
        # 1️⃣ Get Event
        event = session.exec(
            select(Event).where(Event.name == event_name)
        ).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        
        response = []
        # 🔓 CASE 1: OPEN SELL MODE (Sunday / full counter)
        if event.selling_mode == "open_sell":
            # (Optional but recommended) check if user is allowed seller
            allowed = session.exec(
                select(EventSeller).where(
                    EventSeller.event_id == event.id,
                    EventSeller.user_id == user_id
                )
            ).first()
            user = session.get(User, user_id)
            if not allowed and (not user or user.role.lower() != "admin"):
                raise HTTPException(
                    status_code=403,
                    detail="You are not assigned as seller for this event"
                )
            stocks = session.exec(
                select(EventStock, Product)
                .join(Product, Product.id == EventStock.product_id)
                .where(
                    EventStock.event_id == event.id,
                    EventStock.quantity_remaining > 0
                )
            ).all()
            for s, p in stocks:
                response.append({
                    "id": p.id,
                    "name": p.name,
                    "category": p.category,
                    "rate": p.rate,
                    "quantity_remaining": s.quantity_remaining,
                    "is_gift": s.is_gift
                })
        # 🔒 CASE 2: AREA-WISE MODE (DEFAULT)
        else:
            user_areas = session.exec(
                select(UserArea.category)
                .where(UserArea.user_id == user_id)
            ).all()
            if not user_areas:
                return {"event": event.name, "user_id": user_id, "items": []}
            stocks = session.exec(
                select(EventStock, Product)
                .join(Product, Product.id == EventStock.product_id)
                .where(
                    EventStock.event_id == event.id,
                    Product.category.in_(user_areas),
                    EventStock.quantity_remaining > 0
                )
            ).all()
            for s, p in stocks:
                response.append({
                    "id": p.id,
                    "name": p.name,
                    "category": p.category,
                    "rate": p.rate,
                    "quantity_remaining": s.quantity_remaining,
                    "is_gift": s.is_gift
                })
        return {
            "event": event.name,
            "user_id": user_id,
            "items": response
        }


@router.get("/user/sales/daily")
def user_daily_sales(user_id: int):
    with Session(engine) as session:
        sales = session.exec(
            select(Sale)
            .where(Sale.user_id == user_id)
            .order_by(Sale.created_at.desc())
        ).all()
        response = []
        for sale in sales:
            items = session.exec(
                select(SaleItem, Product)
                .join(Product, Product.id == SaleItem.product_id)
                .where(SaleItem.sale_id == sale.id)
            ).all()
            item_list = []
            for si, product in items:
                item_list.append({
                    "name": product.name,
                    "category": product.category,
                    "quantity": si.quantity,
                    "rate": si.rate,
                    "total_price": si.total_price
                })
            response.append({
                "sale_id": sale.id,
                "date_time": sale.created_at,
                "payment_mode": sale.payment_mode,
                "total_amount": sale.total_amount,
                "items": item_list
            })
        return {
            "user_id": user_id,
            "daily_sales": response
        }
@router.get("/user/sales/event")
def user_event_sales(user_id: int, event_name: str | None = None):
    with Session(engine) as session:
        query = select(EventSale).where(EventSale.user_id == user_id)
        if event_name:
            event = session.exec(
                select(Event).where(Event.name == event_name)
            ).first()
            if not event:
                raise HTTPException(status_code=404, detail="Event not found")
            query = query.where(EventSale.event_id == event.id)
        sales = session.exec(
            query.order_by(EventSale.created_at.desc())
        ).all()
        response = []
        for sale in sales:
            items = session.exec(
                select(EventSaleItem, Product)
                .join(Product, Product.id == EventSaleItem.product_id)
                .where(EventSaleItem.event_sale_id == sale.id)
            ).all()
            item_list = []
            for si, product in items:
                item_list.append({
                    "name": product.name,
                    "category": product.category,
                    "quantity": si.quantity,
                    "rate": si.rate,
                    "total_price": si.total_price
                })
            response.append({
                "event_id": sale.event_id,
                "day_number": sale.day_number,
                "date_time": sale.created_at,
                "payment_mode": sale.payment_mode,
                "total_amount": sale.total_amount,
                "items": item_list
            })
        return {
            "user_id": user_id,
            "event_sales": response
        }
@router.get("/user/dashboard")
def user_dashboard(user_id: int):
    with Session(engine) as session:

        # 1️⃣ Today and Month range
        today_start = datetime.combine(date.today(), datetime.min.time())
        yesterday_start = today_start - timedelta(days=1)
        month_start = today_start.replace(day=1)
        today_end = today_start + timedelta(days=1)

        # 2️⃣ User areas
        user_areas = session.exec(
            select(UserArea.category)
            .where(UserArea.user_id == user_id)
        ).all()

        # 3️⃣ Daily sales today
        daily_sales = session.exec(
            select(Sale)
            .where(
                Sale.user_id == user_id,
                Sale.created_at >= today_start,
                Sale.created_at < today_end
            )
        ).all()

        # 4️⃣ Event sales today
        event_sales = session.exec(
            select(EventSale)
            .where(
                EventSale.user_id == user_id,
                EventSale.created_at >= today_start,
                EventSale.created_at < today_end
            )
        ).all()

        # Calculate Month Sales
        month_sales = session.exec(
            select(Sale).where(Sale.user_id == user_id, Sale.created_at >= month_start)
        ).all()
        month_event_sales = session.exec(
            select(EventSale).where(EventSale.user_id == user_id, EventSale.created_at >= month_start)
        ).all()

        # Calculate Yesterday Sales
        yesterday_sales = session.exec(
            select(Sale).where(Sale.user_id == user_id, Sale.created_at >= yesterday_start, Sale.created_at < today_start)
        ).all()
        yesterday_event_sales = session.exec(
            select(EventSale).where(EventSale.user_id == user_id, EventSale.created_at >= yesterday_start, EventSale.created_at < today_start)
        ).all()

        today_revenue = sum(s.total_amount for s in daily_sales + event_sales)
        yesterday_revenue = sum(s.total_amount for s in yesterday_sales + yesterday_event_sales)
        month_revenue = sum(s.total_amount for s in month_sales + month_event_sales)

        # 5️⃣ Active event
        active_event = session.exec(
            select(Event).where(Event.status == "active")
        ).first()

        # 6️⃣ Low stock (main vault)
        low_stock_main = []

        products = session.exec(
            select(Product)
            .where(
                Product.category.in_(user_areas),
                Product.is_active == True
            )
        ).all()

        for product in products:
            batches = session.exec(
                select(ProductBatch.quantity)
                .where(ProductBatch.product_id == product.id)
            ).all()
            total_qty = sum(batches)
            if total_qty <= 5:
                low_stock_main.append({
                    "name": product.name,
                    "category": product.category,
                    "remaining": total_qty
                })
        return {
            "user_id": user_id,
            "today_revenue": today_revenue,
            "yesterday_revenue": yesterday_revenue,
            "month_revenue": month_revenue,
            "daily_sales": len(daily_sales) + len(event_sales),
            "critical_stock": len(low_stock_main),
            "low_stock": len(low_stock_main),
            "active_event": {"id": active_event.id, "name": active_event.name, "mode": active_event.mode if hasattr(active_event, 'mode') else "single-day"} if active_event else None
        }

@router.get("/events/{event_id}/dashboard")
def get_event_dashboard(event_id: int):
    with Session(engine) as session:
        sales = session.exec(
            select(EventSale).where(EventSale.event_id == event_id)
        ).all()
        revenue = sum(s.total_amount for s in sales)
        return {"revenue": revenue}
@router.get("/analytics/users")
def users_analytics():

    with Session(engine) as session:

        users = session.exec(select(User)).all()

        data = []

        for user in users:

            sales = session.exec(
                select(Sale).where(Sale.sold_by == user.name)
            ).all()

            revenue = sum(s.quantity * s.rate for s in sales)

            data.append({
                "name": user.name,
                "role": user.role,
                "sales": len(sales),
                "revenue": revenue
            })

        return data

@router.get("/events/{event_name}/sales/user/{user_id}")
def get_user_event_sales(event_name: str, user_id: int):
    with Session(engine) as session:
        event = session.exec(select(Event).where(Event.name == event_name)).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        
        sales = session.exec(
            select(EventSale).where(
                EventSale.event_id == event.id,
                EventSale.user_id == user_id
            ).order_by(EventSale.created_at.desc())
        ).all()
        
        history = []
        for sale in sales:
            items = session.exec(
                select(EventSaleItem, Product)
                .join(Product, EventSaleItem.product_id == Product.id)
                .where(EventSaleItem.event_sale_id == sale.id)
            ).all()
            
            for item, product in items:
                history.append({
                    "sale_id": sale.id,
                    "date": sale.created_at.isoformat(),
                    "day_number": sale.day_number,
                    "payment_mode": sale.payment_mode,
                    "product_name": product.name,
                    "category": product.category,
                    "quantity": item.quantity,
                    "rate": item.rate,
                    "total_price": item.total_price
                })
        return {
            "event": event.name,
            "user_id": user_id,
            "sales_history": history
        }

@router.get("/sales/daily/monthly")
def get_monthly_daily_sales(month: int = None, year: int = None):
    import datetime
    now = datetime.datetime.now()
    month = month or now.month
    year = year or now.year

    start_date = datetime.datetime(year, month, 1)
    if month == 12:
        end_date = datetime.datetime(year + 1, 1, 1)
    else:
        end_date = datetime.datetime(year, month + 1, 1)
        
    with Session(engine) as session:
        sales = session.exec(
            select(Sale)
            .where(Sale.sale_type == "daily")
            .where(Sale.created_at >= start_date)
            .where(Sale.created_at < end_date)
            .order_by(Sale.created_at.desc())
        ).all()
        
        history = []
        for sale in sales:
            items = session.exec(
                select(SaleItem, Product)
                .join(Product, SaleItem.product_id == Product.id)
                .where(SaleItem.sale_id == sale.id)
            ).all()
            
            for item, product in items:
                history.append({
                    "sale_id": sale.id,
                    "date": sale.created_at.isoformat(),
                    "payment_mode": sale.payment_mode,
                    "product_name": product.name,
                    "category": product.category,
                    "quantity": item.quantity,
                    "rate": item.rate,
                    "total_price": item.total_price
                })
                
        return {
            "month": month,
            "year": year,
            "sales_history": history
        }
