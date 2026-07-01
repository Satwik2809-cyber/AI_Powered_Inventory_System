import io
import os
import re
import math
import shutil
import time

import openpyxl
import pandas as pd
import imagehash
from PIL import Image

from fastapi import APIRouter, HTTPException, UploadFile, File, Depends, Form
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select, col
from datetime import date, datetime

from database import engine
from models import (
    Product, ProductBatch, UserArea, StockHistory, User, Category,
    ProductCreateRequest, ProductRestockRequest, ProductUpdateRequest
)
from auth_and_users import get_current_user

# ─────────────────────────────────────────────
router = APIRouter(tags=["Inventory"])

PRODUCT_IMAGE_DIR = "product_images"

# Image hash cache to avoid re-hashing on every request
_IMAGE_HASH_CACHE: dict = {}

# ─────────────────────────────────────────────
# HELPER FUNCTIONS
# ─────────────────────────────────────────────

def normalize(text: str) -> str:
    return re.sub(r"\s+", " ", text.strip().lower())


def find_product_image(product_name: str) -> str | None:
    if not os.path.exists(PRODUCT_IMAGE_DIR):
        return None
    target = normalize(product_name)
    for file in os.listdir(PRODUCT_IMAGE_DIR):
        name_only = os.path.splitext(file)[0]
        if normalize(name_only) == target:
            return os.path.join(PRODUCT_IMAGE_DIR, file)
    return None


def get_product_by_identity(
    session: Session,
    name: str,
    category: str | None = None,
    rate: float | None = None
):
    name = normalize(name)
    query = select(Product).where(
        Product.is_active == True,  # noqa: E712
        col(Product.name).ilike(f"%{name}%")
    )
    if category:
        query = query.where(Product.category == category)
    if rate:
        query = query.where(Product.rate == rate)

    products = session.exec(query).all()

    if not products:
        raise HTTPException(
            status_code=404,
            detail=f"No product found for '{name}'"
        )
    if len(products) > 1:
        raise HTTPException(
            status_code=409,
            detail={
                "message": "Multiple products found",
                "options": [
                    {"id": p.id, "name": p.name, "rate": p.rate, "category": p.category}
                    for p in products
                ]
            }
        )
    return products[0]


def get_image_hash(image_path: str):
    if image_path in _IMAGE_HASH_CACHE:
        return _IMAGE_HASH_CACHE[image_path]
    try:
        img = Image.open(image_path).convert("RGB")
        h = imagehash.phash(img)
        _IMAGE_HASH_CACHE[image_path] = h
        return h
    except Exception as e:
        print(f"Error hashing image {image_path}: {e}")
        raise e


def match_product_by_image(uploaded_image_path: str):
    uploaded_hash = get_image_hash(uploaded_image_path)
    best_match = None
    min_diff = 999
    with Session(engine) as session:
        products = session.exec(
            select(Product).where(Product.image_path != None)  # noqa: E711
        ).all()
        for product in products:
            try:
                db_hash = get_image_hash(product.image_path)
                diff = uploaded_hash - db_hash
                if diff < min_diff:
                    min_diff = diff
                    best_match = product
            except Exception:
                continue
    if min_diff <= 10:
        return best_match
    return None


# ─────────────────────────────────────────────
# CATEGORIES
# ─────────────────────────────────────────────

@router.get("/categories")
def get_categories():
    with Session(engine) as session:
        categories = session.exec(select(Category).order_by(Category.name)).all()
        return [c.name for c in categories]

@router.post("/categories")
def add_category(name: str = Form(...), current_user: User = Depends(get_current_user)):
    name = name.strip()
    if not name:
        raise HTTPException(400, "Category name cannot be empty")
    
    with Session(engine) as session:
        existing = session.exec(select(Category).where(Category.name == name)).first()
        if existing:
            raise HTTPException(400, "Category already exists")
            
        new_cat = Category(name=name)
        session.add(new_cat)
        session.commit()
        return {"message": "Category created", "name": name}

@router.put("/categories/{old_name}")
def update_category(old_name: str, new_name: str = Form(...), current_user: User = Depends(get_current_user)):
    new_name = new_name.strip()
    if not new_name:
        raise HTTPException(400, "New category name cannot be empty")
        
    with Session(engine) as session:
        cat = session.exec(select(Category).where(Category.name == old_name)).first()
        if not cat:
            raise HTTPException(404, "Category not found")
            
        existing = session.exec(select(Category).where(Category.name == new_name)).first()
        if existing:
            raise HTTPException(400, "Category with that new name already exists")
            
        # Update category
        cat.name = new_name
        
        # Sync old category names in products
        products = session.exec(select(Product).where(Product.category == old_name)).all()
        for p in products:
            p.category = new_name
            
        # Sync old category names in UserArea? 
        # Actually UserArea doesn't exist as a model directly for assignment, users store it as JSON string or UserArea table? 
        # Wait, UserArea IS a model in database? Let me check models.py
        user_areas = session.exec(select(UserArea).where(UserArea.category == old_name)).all()
        for ua in user_areas:
            ua.category = new_name
            
        session.commit()
        return {"message": "Category updated", "old_name": old_name, "new_name": new_name}

@router.delete("/categories/{name}")
def delete_category(name: str, current_user: User = Depends(get_current_user)):
    with Session(engine) as session:
        cat = session.exec(select(Category).where(Category.name == name)).first()
        if not cat:
            raise HTTPException(404, "Category not found")
            
        session.delete(cat)
        
        # Also remove from UserArea so users don't have broken areas
        user_areas = session.exec(select(UserArea).where(UserArea.category == name)).all()
        for ua in user_areas:
            session.delete(ua)
            
        session.commit()
        return {"message": "Category deleted"}


# ─────────────────────────────────────────────
# PRODUCT ROUTES (MAIN VAULT)
# ─────────────────────────────────────────────

@router.post("/products")
def add_product(
    product_data: ProductCreateRequest,
    current_user: User = Depends(get_current_user)
):
    with Session(engine) as session:
        product = Product(
            name=product_data.name,
            category=product_data.category,
            rate=product_data.rate,
            is_active=True
        )
        session.add(product)
        session.commit()
        session.refresh(product)

        if product_data.quantity > 0:
            batch = ProductBatch(
                product_id=product.id,
                quantity=product_data.quantity,
                manufacturing_date=product_data.manufacturing_date,
                expiry_date=product_data.expiry_date,
                last_restocked_at=datetime.utcnow()
            )
            history = StockHistory(
                product_id=product.id,
                quantity=product_data.quantity,
                manufacturing_date=str(product_data.manufacturing_date) if product_data.manufacturing_date else None,
                expiry_date=str(product_data.expiry_date) if product_data.expiry_date else None,
                restocked_by=current_user.username,
                created_at=datetime.utcnow()
            )
            session.add(batch)
            session.add(history)
            session.commit()

        return product


@router.get("/products")
def list_products():
    with Session(engine) as session:
        products = session.exec(
            select(Product).where(Product.is_active == True)  # noqa: E712
        ).all()

        results = []
        for p in products:
            stock_data = session.exec(
                select(ProductBatch.quantity, ProductBatch.manufacturing_date, ProductBatch.expiry_date)
                .where(ProductBatch.product_id == p.id)
                .where(ProductBatch.quantity > 0)
                .order_by(col(ProductBatch.expiry_date).asc())
            ).all()

            total_qty = sum(item.quantity for item in stock_data)

            p_dict = p.dict()

            # Sanitize float values to avoid JSON errors
            if p_dict.get("rate") is not None and (math.isnan(p_dict["rate"]) or math.isinf(p_dict["rate"])):
                p_dict["rate"] = 0.0

            p_dict["quantity"] = total_qty if total_qty else 0

            if stock_data:
                p_dict["manufacturing_date"] = stock_data[0].manufacturing_date
                p_dict["expiry_date"] = stock_data[0].expiry_date
            else:
                p_dict["manufacturing_date"] = None
                p_dict["expiry_date"] = None

            results.append(p_dict)

        return results


@router.put("/products/{product_id}")
def update_product(product_id: int, update_data: ProductUpdateRequest):
    with Session(engine) as session:
        product = session.get(Product, product_id)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

        product.name = update_data.name
        product.category = update_data.category
        product.rate = update_data.rate
        session.commit()

        return {"message": "Product updated successfully"}


@router.delete("/products")
def deactivate_product(name: str, category: str, rate: float):
    with Session(engine) as session:
        product = get_product_by_identity(session, name, category, rate)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

        product.is_active = False
        session.commit()
        return {"message": "Product deactivated"}


# ─────────────────────────────────────────────
# STOCK MANAGEMENT
# ─────────────────────────────────────────────

@router.post("/products/restock")
def restock_product(
    restock_data: ProductRestockRequest,
    current_user: User = Depends(get_current_user)
):
    if restock_data.expiry_date and restock_data.manufacturing_date:
        if restock_data.expiry_date <= restock_data.manufacturing_date:
            raise HTTPException(
                status_code=400,
                detail="Expiry date must be after manufacturing date"
            )

    with Session(engine) as session:
        product = get_product_by_identity(session, restock_data.name, restock_data.category, restock_data.rate)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

        batch = ProductBatch(
            product_id=product.id,
            quantity=restock_data.quantity,
            manufacturing_date=restock_data.manufacturing_date,
            expiry_date=restock_data.expiry_date,
            last_restocked_at=datetime.utcnow()
        )

        history = StockHistory(
            product_id=product.id,
            quantity=restock_data.quantity,
            manufacturing_date=str(restock_data.manufacturing_date) if restock_data.manufacturing_date else None,
            expiry_date=str(restock_data.expiry_date) if restock_data.expiry_date else None,
            restocked_by=current_user.username,
            created_at=datetime.utcnow()
        )

        session.add(batch)
        session.add(history)
        session.commit()
        return {"message": "Stock added successfully"}


@router.get("/products/stock-history")
def get_stock_history(name: str, category: str):
    with Session(engine) as session:
        product = session.exec(
            select(Product).where(
                Product.name == name,
                Product.category == category
            )
        ).first()

        if not product:
            raise HTTPException(404, "Product not found")

        records = session.exec(
            select(StockHistory)
            .where(StockHistory.product_id == product.id)
            .order_by(StockHistory.created_at.desc())
        ).all()

        return records


@router.get("/products/stock")
def get_product_stock(name: str, category: str, rate: float):
    with Session(engine) as session:
        product = get_product_by_identity(session, name, category, rate)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

        return session.exec(
            select(ProductBatch).where(ProductBatch.product_id == product.id)
        ).all()


@router.get("/products/quantity")
def get_total_quantity(name: str, category: str, rate: float):
    with Session(engine) as session:
        product = get_product_by_identity(session, name, category, rate)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

        batches = session.exec(
            select(ProductBatch).where(ProductBatch.product_id == product.id)
        ).all()

        total = sum(batch.quantity for batch in batches)
        return {
            "product": name,
            "total_quantity": total
        }


@router.get("/products/user-view")
def products_for_user(user_id: int):
    with Session(engine) as session:
        user_areas = session.exec(
            select(UserArea.category)
            .where(UserArea.user_id == user_id)
        ).all()
        return session.exec(
            select(Product)
            .where(
                Product.category.in_(user_areas),
                Product.is_active == True  # noqa: E712
            )
        ).all()


@router.get("/products/stock/user")
def get_user_main_stock(user_id: int):
    with Session(engine) as session:
        user_areas = session.exec(
            select(UserArea.category)
            .where(UserArea.user_id == user_id)
        ).all()
        if not user_areas:
            return {"items": []}

        products = session.exec(
            select(Product)
            .where(
                Product.category.in_(user_areas),
                Product.is_active == True  # noqa: E712
            )
        ).all()
        response = []
        for product in products:
            batches = session.exec(
                select(ProductBatch.quantity)
                .where(ProductBatch.product_id == product.id)
            ).all()
            total_qty = sum(q for (q,) in batches)
            response.append({
                "name": product.name,
                "category": product.category,
                "rate": product.rate,
                "total_quantity": total_qty,
                "low_stock_alert": total_qty <= 5
            })
        return {
            "user_id": user_id,
            "items": response
        }


# ─────────────────────────────────────────────
# EXCEL IMPORT / EXPORT
# ─────────────────────────────────────────────

@router.post("/products/import-excel")
async def import_products_from_excel(
    file: UploadFile = File(...),
    mode: str = Form("registry"),
    current_user: User = Depends(get_current_user)
):
    print(f"DEBUG: Processing file: {file.filename}")
    if not file.filename.endswith(".xlsx"):
        print(f"DEBUG: Invalid extension: {file.filename}")
        raise HTTPException(400, "Only Excel files allowed")

    contents = await file.read()
    try:
        df = pd.read_excel(io.BytesIO(contents))
    except Exception as e:
        print(f"DEBUG: Pandas read failed: {e}")
        raise HTTPException(400, "Could not parse Excel file. Ensure valid format.")

    print(f"DEBUG: Columns found: {df.columns.tolist()}")

    df.columns = df.columns.str.strip().str.lower()

    rename_map = {
        "item name": "name",
        "item": "name",
        "product": "name",
        "categories": "category",
        "cat": "category",
        "price": "rate",
        "mrp": "rate",
        "cost": "rate"
    }
    df.rename(columns=rename_map, inplace=True)

    print(f"DEBUG: Normalized columns: {df.columns.tolist()}")

    if mode == "registry":
        required_cols = {"name", "category", "rate"}
        if not required_cols.issubset(df.columns):
            raise HTTPException(400, f"Excel must contain at least: {required_cols}")
    else:
        # restock mode
        required_cols = {"name", "quantity"}
        if not required_cols.issubset(df.columns):
            raise HTTPException(400, f"Excel must contain at least: {required_cols}")

    if "rate" in df.columns:
        df["rate"] = df["rate"].fillna(0.0)
    if "category" in df.columns:
        df["category"] = df["category"].fillna("Uncategorized")
    if "quantity" in df.columns:
        df["quantity"] = df["quantity"].fillna(0)

    created_products = []
    restocked_products = []

    with Session(engine) as session:
        for _, row in df.iterrows():
            name = str(row["name"]).strip()
            
            # For registry mode, category and rate are guaranteed by required_cols. For restock they might be missing.
            category = str(row.get("category", "Uncategorized")).strip() if pd.notna(row.get("category", "Uncategorized")) else "Uncategorized"
            rate = float(row.get("rate", 0.0)) if pd.notna(row.get("rate", 0.0)) else 0.0

            quantity = 0
            if mode == "restock" and "quantity" in df.columns:
                try:
                    q = int(row["quantity"])
                    if not pd.isna(q):
                        quantity = q
                except Exception:
                    quantity = 0

            mfg = None
            exp = None
            if "manufacturing_date" in df.columns and "expiry_date" in df.columns:
                try:
                    m = pd.to_datetime(row["manufacturing_date"]).date()
                    e = pd.to_datetime(row["expiry_date"]).date()
                    if not pd.isna(m) and not pd.isna(e) and e >= m:
                        mfg = m
                        exp = e
                except Exception:
                    pass

            product = session.exec(
                select(Product).where(
                    Product.name == name,
                    Product.category == category
                )
            ).first()

            if not product:
                product = Product(
                    name=name,
                    category=category,
                    rate=rate,
                    is_active=True,
                    image_path=find_product_image(name)
                )
                session.add(product)
                session.commit()
                session.refresh(product)
                created_products.append(name)

            if quantity > 0:
                batch = ProductBatch(
                    product_id=product.id,
                    quantity=quantity,
                    manufacturing_date=mfg,
                    expiry_date=exp,
                    last_restocked_at=datetime.utcnow()
                )
                history = StockHistory(
                    product_id=product.id,
                    quantity=quantity,
                    manufacturing_date=str(mfg) if mfg else None,
                    expiry_date=str(exp) if exp else None,
                    restocked_by=current_user.username
                )
                session.add(batch)
                session.add(history)
                restocked_products.append(name)

        session.commit()

    return {
        "created_products": created_products,
        "restocked_products": restocked_products,
        "total_rows_processed": len(df)
    }


@router.get("/products/export-excel")
def export_products_excel():
    with Session(engine) as session:
        products = session.exec(select(Product)).all()
        products = sorted(products, key=lambda p: (p.category or "Uncategorized", p.name))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        headers = ["Name", "Category", "Rate", "Total Quantity", "Restock Date", "Item Restock Quantity", "Restocked By", "Active"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)

        current_cat = None
        for p in products:
            cat = p.category or "Uncategorized"
            if cat != current_cat:
                current_cat = cat
                ws.append([f"--- {cat.upper()} ---"] + [""] * 7)
                for col_idx in range(1, len(headers) + 1):
                    ws.cell(row=ws.max_row, column=col_idx).font = openpyxl.styles.Font(bold=True, italic=True)
                    ws.cell(row=ws.max_row, column=col_idx).fill = openpyxl.styles.PatternFill(
                        start_color="EFEFEF", end_color="EFEFEF", fill_type="solid"
                    )

            qty = session.exec(
                select(ProductBatch.quantity)
                .where(ProductBatch.product_id == p.id)
            ).all()
            total_qty = sum(qty)

            latest_history = session.exec(
                select(StockHistory)
                .where(StockHistory.product_id == p.id)
                .order_by(StockHistory.created_at.desc())
            ).first()

            restock_date = latest_history.created_at.strftime("%Y-%m-%d %H:%M:%S") if latest_history else "N/A"
            restock_qty = latest_history.quantity if latest_history else 0
            restocked_by = latest_history.restocked_by if latest_history else "N/A"

            ws.append([
                p.name,
                p.category,
                p.rate,
                total_qty,
                restock_date,
                restock_qty,
                restocked_by,
                "Yes" if p.is_active else "No"
            ])

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column].width = max_length + 2

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"products_export_{timestamp}.xlsx"

        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )


# ─────────────────────────────────────────────
# IMAGE ROUTES
# ─────────────────────────────────────────────

@router.post("/products/scan-image")
def scan_product_image(file: UploadFile = File(...)):
    temp_path = f"temp_{file.filename}"

    with open(temp_path, "wb") as f:
        f.write(file.file.read())

    product = match_product_by_image(temp_path)
    os.remove(temp_path)

    if not product:
        raise HTTPException(
            status_code=404,
            detail="Product not recognized"
        )

    return {
        "id": product.id,
        "name": product.name,
        "category": product.category,
        "rate": product.rate,
        "image": product.image_path
    }


@router.post("/products/{product_id}/image")
def upload_product_image(product_id: int, file: UploadFile = File(...)):
    if not os.path.exists(PRODUCT_IMAGE_DIR):
        os.makedirs(PRODUCT_IMAGE_DIR)

    with Session(engine) as session:
        product = session.get(Product, product_id)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

        ext = os.path.splitext(file.filename)[1]
        if not ext:
            ext = ".jpg"

        timestamp = int(time.time())
        safe_name = normalize(product.name).replace(" ", "_").replace("/", "_")
        filename = f"{safe_name}_{timestamp}{ext}"
        filepath = os.path.join(PRODUCT_IMAGE_DIR, filename)

        with open(filepath, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        product.image_path = filepath
        session.commit()

        if filepath in _IMAGE_HASH_CACHE:
            del _IMAGE_HASH_CACHE[filepath]

        return {"message": "Image uploaded successfully", "image_path": filepath}


@router.get("/products/{product_id}/images")
def get_product_images(product_id: int):
    with Session(engine) as session:
        product = session.get(Product, product_id)
        if not product:
            raise HTTPException(status_code=404, detail="Product not found")

        if not os.path.exists(PRODUCT_IMAGE_DIR):
            return {"images": []}

        target = normalize(product.name)
        safe_target = target.replace(" ", "_").replace("/", "_")
        images = []
        for f in os.listdir(PRODUCT_IMAGE_DIR):
            name_only = os.path.splitext(f)[0]
            normalized_file = normalize(name_only)
            if normalized_file == target or normalized_file.startswith(target + "_") or normalized_file.startswith(safe_target + "_"):
                images.append(f"/product_images/{f}")

        try:
            images.sort(key=lambda x: os.path.getmtime(os.path.join(".", x.lstrip("/"))), reverse=True)
        except Exception:
            pass
        return {"images": images}


@router.get("/products/all-images")
def get_all_product_images():
    if not os.path.exists(PRODUCT_IMAGE_DIR):
        return {"images": []}

    images = []
    for f in os.listdir(PRODUCT_IMAGE_DIR):
        if f.lower().endswith(('.png', '.jpg', '.jpeg', '.webp', '.gif')):
            images.append({
                "url": f"/product_images/{f}",
                "name": f
            })

    try:
        images.sort(key=lambda x: os.path.getmtime(os.path.join(".", x["url"].lstrip("/"))), reverse=True)
    except Exception:
        pass

    return {"images": images}


@router.delete("/products/all-images/{filename}")
def delete_library_image(filename: str):
    filepath = os.path.join(PRODUCT_IMAGE_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(404, "Image not found")

    try:
        os.remove(filepath)
        if filepath in _IMAGE_HASH_CACHE:
            del _IMAGE_HASH_CACHE[filepath]
        return {"message": "Image deleted successfully"}
    except Exception as e:
        raise HTTPException(500, f"Delete failed: {str(e)}")


@router.post("/products/all-images/update")
def update_library_image(filename: str = Form(...), file: UploadFile = File(...)):
    filepath = os.path.join(PRODUCT_IMAGE_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(404, "Image not found")

    try:
        with open(filepath, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        if filepath in _IMAGE_HASH_CACHE:
            del _IMAGE_HASH_CACHE[filepath]

        return {"message": "Image updated successfully"}
    except Exception as e:
        raise HTTPException(500, f"Update failed: {str(e)}")
